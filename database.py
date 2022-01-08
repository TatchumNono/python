from openpyxl import load_workbook
import pandas as pd
import numpy

filename = "C:\\Users\\TatchumNono\\Documents\\employeedata.xlsx"

wb = load_workbook(filename)

sheet = wb['Feuil1']

# array to store the emails to be modified extracted from the .xlsx file
emails = []

# array containing the modified emails
modified_emails = []

# function to extract the emails from the .xlsx file and store them in emails[]


def extract_emails():
    for x in range(sheet._current_row):
        if(x == 0):
            #print('Jumping the headers')
            print(' ')
        else:
            mail = sheet.cell(row=x + 1, column=2).value
            emails.append(mail)

# function to modify the emails stores in emails[] ans store them in modified_emails[]


def modify_emails():
    length = len(emails)
    for x in range(length):
        mail = emails[x].split("@")
        result = mail[0] + '@handsinhands.org'
        modified_emails.append(result)

# function to modify the .xlsx file using the modified emails stored in modified_emails[]


def modify_xlsx():
    for x in range(sheet._current_row):
        if(x == 0):
            #print('Jumping the headers')
            print(' ')
        else:
            sheet.cell(row=x + 1, column=2).value = modified_emails[x-1]
            wb.save(filename)

# function to modify the .csv file using the modified emails stored in modified_emails[]


def modify_csv():
    dx = pd.read_csv(r'C:\Users\TatchumNono\Documents\employeedata.csv')
    size = len(modified_emails)
    for x in range(size):
        dx.loc[x, 'email'] = modified_emails[x]
        dx.to_csv('C:\\Users\TatchumNono\Documents\employeedata.csv', index=False)

# function to display the .xlsx file


def display_xlsx_file():
    df = pd.read_excel(r'C:\Users\TatchumNono\Documents\employeedata.xlsx')
    print(df)

# function to display the .csv file


def display_csv_file():
    df = pd.read_csv(r'C:\Users\TatchumNono\Documents\employeedata.csv')
    print(df)


print("Extracting the emails to be modified")

extract_emails()

print("Modifying the emails...")

modify_emails()

print(' ')

print("Modifying both the .xlsx and .csv files and updating thier emails")

print(' ')

print("Before modification")

print(' ')

display_xlsx_file()

print(' ')

display_csv_file()

print(' ')

print('Modifying...')

modify_xlsx()

modify_csv()

print(' ')

print("After modification")

print(' ')

display_xlsx_file()

print(' ')

display_csv_file()
